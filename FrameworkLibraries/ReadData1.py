import openpyxl

Workbook=openpyxl.load_workbook("../../TestData/TestData2.xlsx")

def fetch_number_of_rows(Sheetname):
   Sheet1 = Workbook[Sheetname]
   return Sheet1.max_row

def fetch_cell_data(Sheetname,row,cell):
    Sheet1 = Workbook[Sheetname]
    cell = Sheet1.cell(int(row),int(cell))
    return cell.value

print (fetch_number_of_rows("Sheet1"))
# print(fetch_cell_data("Sheet1",1,1))